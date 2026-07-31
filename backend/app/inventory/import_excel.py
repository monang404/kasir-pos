from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import text
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.auth.require_role import RequireModule
from app.auth.session import get_current_user
from app.database import get_db
from app.activity_log.logger import log_action

router = APIRouter(prefix="/inventory/import", tags=["inventory"])
check_inventory_access = RequireModule("inventory")

@router.post("/", dependencies=[Depends(check_inventory_access)])
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="File harus berformat .xlsx")
        
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        
        # Cari sheet 'produk', fallback ke sheet pertama
        sheet_name = 'produk' if 'produk' in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        # Cari indeks header
        header_row = None
        headers = {}
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any(isinstance(cell, str) and cell.strip().lower() == 'kode' for cell in row):
                header_row = row_idx
                for col_idx, cell in enumerate(row):
                    if isinstance(cell, str):
                        headers[cell.strip().lower()] = col_idx
                break
                
        if not header_row or 'kode' not in headers or 'nama' not in headers or 'harga_beli' not in headers or 'harga_jual' not in headers:
            raise HTTPException(status_code=400, detail="Format header tidak valid. Wajib ada: kode, nama, harga_beli, harga_jual")

        berhasil = 0
        dilewati = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not any(row): continue # Skip empty row
            
            try:
                kode_raw = row[headers['kode']]
                nama_raw = row[headers['nama']]
                hb_raw = row[headers['harga_beli']]
                hj_raw = row[headers['harga_jual']]
                
                # Cek kolom wajib isi
                if not kode_raw or not nama_raw:
                    dilewati += 1
                    continue
                    
                kode = str(kode_raw).strip().upper()
                nama = str(nama_raw).strip()
                
                if not kode or not nama:
                    dilewati += 1
                    continue
                    
                hb = float(hb_raw) if hb_raw is not None else 0.0
                hj = float(hj_raw) if hj_raw is not None else 0.0
                
                if hj <= 0:
                    errors.append(f"Baris {row_idx}: Harga Jual <= 0")
                    dilewati += 1
                    continue
                    
                if hj < hb:
                    errors.append(f"Baris {row_idx}: Harga Jual < Harga Beli")
                    dilewati += 1
                    continue
                    
                # Insert with savepoint
                try:
                    with db.begin_nested():
                        db.execute(
                            text("""
                                INSERT INTO produk (kode, nama, harga_beli, harga_jual)
                                VALUES (:kode, :nama, :hb, :hj)
                            """),
                            {"kode": kode, "nama": nama, "hb": hb, "hj": hj}
                        )
                except IntegrityError:
                    errors.append(f"Baris {row_idx}: Kode '{kode}' sudah ada atau melanggar constraint")
                    dilewati += 1
                    continue
                
                berhasil += 1
                
            except Exception as e:
                import logging
                logging.exception(e)
                errors.append(f"Baris {row_idx}: Error parsing data. Silakan cek format file.")
                dilewati += 1
                continue
                
        # Activity log
        if berhasil > 0:
            log_action(db, user, 'IMPORT', 'inventory', 'EXCEL', f"Import Excel berhasil {berhasil} produk")
            
        db.commit()
        
        return {
            "message": "Import selesai",
            "ringkasan": {
                "berhasil": berhasil,
                "dilewati": dilewati,
                "errors": errors[:10]
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.exception(e)
        raise HTTPException(status_code=500, detail="Gagal memproses file. Silakan coba lagi.")
