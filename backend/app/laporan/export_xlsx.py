import io

from fastapi import APIRouter, Depends, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.auth.require_role import RequireModule
from app.database import get_db
from app.laporan.ringkasan_transaksi_produk import build_date_filter

router = APIRouter(prefix="/laporan", tags=["laporan"])
check_access = RequireModule("laporan")

# STYLING CONSTANTS (Dark Theme Repilca)
COLOR_NAVY = "0a0a2a"
COLOR_NAVY_LIGHT = "11113a"
COLOR_NAVY_LIGHTER = "1e1e4a"
COLOR_TEXT = "e2e8f0"
COLOR_HEADER_BG = "4f46e5"

FONT_TITLE = Font(name='Segoe UI', size=14, bold=True, color=COLOR_TEXT)
FONT_HEADER = Font(name='Segoe UI', size=11, bold=True, color=COLOR_TEXT)
FONT_CELL = Font(name='Segoe UI', size=11, color=COLOR_TEXT)

FILL_BG_MAIN = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
FILL_BG_ALT = PatternFill(start_color=COLOR_NAVY_LIGHT, end_color=COLOR_NAVY_LIGHT, fill_type="solid")
FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
FILL_TITLE = PatternFill(start_color=COLOR_NAVY_LIGHTER, end_color=COLOR_NAVY_LIGHTER, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='2d2d5f'), 
    right=Side(style='thin', color='2d2d5f'), 
    top=Side(style='thin', color='2d2d5f'), 
    bottom=Side(style='thin', color='2d2d5f')
)

def apply_base_style(ws, max_row, max_col):
    """Mengisi background default agar mirip dark mode app"""
    for r in range(1, max_row + 5):
        for c in range(1, max_col + 5):
            cell = ws.cell(row=r, column=c)
            if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                cell.fill = FILL_BG_MAIN
                cell.font = FONT_CELL

def create_excel_file(title: str, headers: list, data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        # Auto width approx
        ws.column_dimensions[c.column_letter].width = max(15, len(h) + 5)

    # Data
    row_idx = 4
    for row_data in data:
        fill_color = FILL_BG_ALT if row_idx % 2 == 0 else FILL_BG_MAIN
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_CELL
            c.fill = fill_color
            c.border = THIN_BORDER
        row_idx += 1

    apply_base_style(ws, row_idx, len(headers))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

@router.get("/export/{tab}", dependencies=[Depends(check_access)])
def export_laporan_excel(
    tab: str,
    mode: str = Query("bulan"),
    bulan: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db)
):
    where_clause, params = build_date_filter(mode, bulan, start_date, end_date, "tanggal")
    
    headers = []
    data = []
    title = f"Laporan {tab.capitalize()}"

    if tab == "transaksi":
        headers = ["Kode", "Tanggal", "Pelanggan", "Kasir", "Metode", "Total", "Profit"]
        rows = db.execute(text(f"""
            SELECT t.kode, t.tanggal, p.nama as pel, t.kasir_nama, t.metode_bayar, t.total, t.profit
            FROM transaksi t LEFT JOIN pelanggan p ON t.pelanggan_id = p.id
            WHERE {where_clause} ORDER BY t.tanggal DESC
        """), params).fetchall()
        for r in rows:
            data.append([r.kode, str(r.tanggal)[:10], r.pel or "Umum", r.kasir_nama, r.metode_bayar, r.total, r.profit])

    elif tab == "produk":
        where_clause_p, params_p = build_date_filter(mode, bulan, start_date, end_date, "t.tanggal")
        headers = ["Kode", "Nama Produk", "Qty Terjual", "Omzet", "Profit"]
        rows = db.execute(text(f"""
            SELECT p.kode, p.nama, SUM(td.qty) as q, SUM((td.harga_jual+td.harga_tinta)*td.qty) as o, SUM(((td.harga_jual+td.harga_tinta)-td.harga_beli)*td.qty) as pr
            FROM transaksi_detail td JOIN transaksi t ON td.transaksi_id = t.id JOIN produk p ON td.produk_id = p.id
            WHERE {where_clause_p} AND td.is_bonus = 0 GROUP BY p.id, p.kode, p.nama ORDER BY q DESC
        """), params_p).fetchall()
        for r in rows:
            data.append([r.kode, r.nama, r.q, r.o, r.pr])

    elif tab == "pelanggan":
        where_clause_p, params_p = build_date_filter(mode, bulan, start_date, end_date, "t.tanggal")
        headers = ["Nama", "No HP", "Jml Trx", "Total Belanja", "Profit"]
        rows = db.execute(text(f"""
            SELECT p.nama, p.no_hp, COUNT(t.id) as c, SUM(t.total) as tb, SUM(t.profit) as pr
            FROM transaksi t JOIN pelanggan p ON t.pelanggan_id = p.id
            WHERE {where_clause_p} GROUP BY p.id, p.nama, p.no_hp ORDER BY tb DESC
        """), params_p).fetchall()
        for r in rows:
            data.append([r.nama, r.no_hp or "-", r.c, r.tb, r.pr])

    elif tab == "pengeluaran":
        headers = ["Kategori", "Total"]
        rows = db.execute(text(f"""
            SELECT kategori, SUM(jumlah) as t FROM pengeluaran WHERE {where_clause} GROUP BY kategori ORDER BY t DESC
        """), params).fetchall()
        for r in rows:
            data.append([r.kategori, r.t])

    elif tab == "stok":
        headers = ["Kode", "Nama", "Qty Sisa", "Harga Beli (HPP)", "Valuasi"]
        rows = db.execute(text("""
            SELECT p.kode, p.nama, COALESCE(SUM(pb.qty_sisa),0) as q, p.harga_beli, COALESCE(SUM(pb.qty_sisa*pb.harga_beli),0) as v
            FROM produk p LEFT JOIN produk_batch pb ON p.id = pb.produk_id
            GROUP BY p.id, p.kode, p.nama, p.harga_beli ORDER BY p.kode
        """)).fetchall()
        title = "Laporan Stok (Snapshot Real-time)"
        for r in rows:
            data.append([r.kode, r.nama, r.q, r.harga_beli, r.v])
            
    else:
        return Response(content="Invalid tab", status_code=400)

    file_bytes = create_excel_file(title, headers, data)
    
    filename = f"laporan_{tab}.xlsx"
    return Response(
        content=file_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
